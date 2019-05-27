import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_spreadsheet(filename):
	wb = xl.load_workbook(filename)
	sheet = wb['Sheet1']



	for row in range(2,sheet.max_row+1):
		cell = sheet.cell(row,3)
		correct = cell.value * 0.9
		correct_cell = sheet.cell(row,4)
		correct_cell.value = correct


	values =Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
	bar = BarChart()
	bar.add_data(values)
	sheet.add_chart(bar,"E2")

	wb.save(filename) 


process_spreadsheet("transactions.xlsx")